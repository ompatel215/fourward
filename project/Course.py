class CourseScheduler:
    def __init__(self):
        # Dictionary to store courses and their prerequisites
        self.graph = {}

    def add_course(self, course):
        """Add a course to the graph."""
        if course is not None and course.strip():  # Check if course is not None and not empty
            if course not in self.graph:
                self.graph[course] = {
                    "prerequisites": [],
                    "credits": "",
                    "description": ""
                }

    def add_prerequisite(self, course, prerequisite):
        """Add a prerequisite relationship (prerequisite -> course)."""
        if course not in self.graph:
            self.add_course(course)
        if prerequisite not in self.graph:
            self.add_course(prerequisite)
        self.graph[course].append(prerequisite)

    def get_prerequisites(self, course, visited=None):
        """Get all prerequisites for a given course."""
        if visited is None:
            visited = set()
        if course not in self.graph:
            return set()
        prerequisites = set()
        course_data = self.graph[course]
        print(f"Course data for {course}: {course_data}")
        for prereq in course_data["prerequisites"]:
            if prereq not in visited:
                visited.add(prereq)
                prerequisites.add(prereq)
                prerequisites.update(self.get_prerequisites(prereq, visited))
        return prerequisites

    def display_courses(self):
        """Display all courses and their prerequisites."""
        for course in self.graph:
            prerequisites = self.get_prerequisites(course)
            print(f"Course: {course}, Prerequisites: {prerequisites}")

    def load_courses_from_hardcoded_data(self):
        """Load courses and prerequisites from hardcoded data."""
        # Example hardcoded data
        data = [
            ("Math101", ""),
            ("Math102", "Math101"),
            ("CS101", ""),
            ("CS102", "CS101, Math101")
        ]

        for course, prerequisites in data:
            self.add_course(course)
            if prerequisites:  # Check if prerequisites is not None
                for prerequisite in prerequisites.split(','):
                    self.add_prerequisite(course, prerequisite.strip())

    def load_courses_from_excel(self, file_path):
        """Load courses, prerequisites, credits, and descriptions from an Excel file."""
        from openpyxl import load_workbook
        workbook = load_workbook(filename=file_path)
        sheet = workbook.active

        for row in sheet.iter_rows(min_row=2, values_only=True):  # Assuming the first row is headers
            course, prerequisites, credits, description = row
            self.add_course(course)
            self.graph[course] = {
                "prerequisites": [],
                "credits": credits,
                "description": description
            }
            if prerequisites:  # Check if prerequisites is not None
                for prerequisite in prerequisites.split(','):
                    self.graph[course]["prerequisites"].append(prerequisite.strip())

        # Debug print to verify graph structure
        print("Graph structure after loading:", self.graph)

    def delete_course(self, course):
        """Delete a course from the graph."""
        if course in self.graph:
            del self.graph[course]
            # Remove this course from any prerequisites lists
            for c in self.graph:
                if course in self.graph[c]["prerequisites"]:
                    self.graph[c]["prerequisites"].remove(course)

    def get_recommended_courses(self, completed_courses):
        """Get recommended courses based on completed courses."""
        recommended = set()
        
        # First, get all prerequisites of completed courses
        prerequisites_to_exclude = set()
        for completed in completed_courses:
            if completed in self.graph:
                prerequisites_to_exclude.update(self.get_prerequisites(completed))
        
        for course in self.graph:
            # Skip if course is already completed or is a prerequisite of completed courses
            if course not in completed_courses and course not in prerequisites_to_exclude:
                # Check if all prerequisites are met
                prerequisites = self.graph[course]["prerequisites"]
                if all(prereq in completed_courses for prereq in prerequisites):
                    recommended.add(course)
            
        return list(recommended)